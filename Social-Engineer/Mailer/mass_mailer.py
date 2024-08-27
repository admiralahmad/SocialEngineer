"""
Program Name: Mailer
Description: This script sends a mass email to recipients.
Author: Ahmad Tijjani Saidu
Date: 17/04/2024
Input: Email template(optional), attachments(optional), excel sheet containing recipient list
Output: Email Sent Successfully
Note: Some parts of this code were developed with the assistance of an AI Chatbot.

How to run - command line
python mass_mailer.py
"""

import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from jinja2 import Template
import openpyxl
from dotenv import load_dotenv

load_dotenv()  # Load environment variables

def send_email(sender_email, subject, message, recipients, template_path=None, filename=None, filetype=None):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(recipients)
    msg['Subject'] = subject

    # Load the email message body from template if exists
    if template_path:
        with open(template_path, 'r') as file:
            message = file.read()
    
    msg.attach(MIMEText(message, 'plain'))
    msg.attach(MIMEText(f"<html><body>{message}</body></html>", 'html'))

    if filename and filetype:
        attach_file(msg, filename, filetype)
      
    # SMTP Server Setup (go to .env file for smtp server configuration)
    try:
        with smtplib.SMTP(os.getenv('SMTP_SERVER'), os.getenv("SMTP_PORT")) as server: 
            server.starttls()
            server.login(os.getenv("USER_NAME"), os.getenv("PASSWORD"))
            server.sendmail(sender_email, recipients, msg.as_string())
            print("Email sent successfully!")
    except smtplib.SMTPException as e:
        print(f"SMTP error occurred: {str(e)}")
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")
    
# File Attachment Function
def attach_file(msg, filename, filetype):
    try:
        att_name = os.path.basename(filename)
        if filetype == 'pdf':
            with open(filename, 'rb') as f:
                att = MIMEApplication(f.read(), _subtype="pdf")
                att.add_header('Content-Disposition', 'attachment', filename=att_name)
                msg.attach(att)
        elif filetype == 'img':
            with open(filename, 'rb') as img_file:
                att = MIMEImage(img_file.read())
        elif filetype == 'word':
                with open(filename, 'rb') as f:
                    att = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.wordprocessingml.document")
        elif filetype == 'excel':
                with open(filename, 'rb') as f:
                    att = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")      
        else:
            print("Unsupported file type provided.")
            return
        
        att.add_header('Content-Disposition', 'attachment', filename=att_name)
        msg.attach(att)
    except FileNotFoundError:
        print(f"Error: The file {filename} was not found.")
    except Exception as e:
        print ({f"An error occured: {str(e)}"})    

# Read Email Addresses From Excel Sheet
def read_email_addresses_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

     # Read email addresses from column A, starting from row 2
    email_addresses = [cell.value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row]
    wb.close()
    return email_addresses


# Main Function
def mass_mailer():
    sender_email = input("Enter sender email: ")
    subject = input("Enter email subject: ")
    generate_from_template = input("Generate email from template? (y/n): ")
    template_path = input("Enter path to HTML template: ") if generate_from_template.lower() == 'y' else None

    message = input("Enter email message (type 'END' to finish): ")
    lines = [] #adding line input
    while True:
        line = input("Next Line: ")
        if line == "END":
            break
        lines.append(line)
    message = "\n".join(lines)
    recipients = read_email_addresses_from_excel(input("Enter path to Excel file containing email addresses: "))

    if input("Do you want to attach a file? (y/n) ") == "y":
        filename = input("Enter the path of the file you want to attach: ")
        filetype = input("What type (pdf, img, word, excel): ")
    else:
        filename = filetype = None

    send_email(sender_email, subject, message, recipients, template_path, filename, filetype)
    

if __name__ == "__main__":
    mass_mailer()
